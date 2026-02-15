from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

from tests.helpers import make_synthetic_yxc, write_ome_tiff


def _run_cli(cmd: list[str]) -> subprocess.CompletedProcess[str]:
    return subprocess.run(cmd, capture_output=True, text=True, check=False)


def test_cli_preview_only_blocks_excel_for_student(tmp_path: Path) -> None:
    img = make_synthetic_yxc(n_nuclei=8)
    inp = tmp_path / "img.ome.tif"
    write_ome_tiff(inp, img, pixel_size_um=0.5)

    out_root = tmp_path / "runs"
    channel_map = json.dumps({"NUCLEUS": 0, "MARKER_1": 1})

    cmd = [
        sys.executable,
        "-m",
        "protocolquant.cli",
        "run",
        "--protocol",
        "protocols/nuclei_count_intensity_2d_v1.yaml",
        "--input",
        str(inp),
        "--role",
        "student",
        "--policy",
        "configs/lab_policy.yaml",
        "--out",
        str(out_root),
        "--preview-count",
        "1",
        "--channel-map",
        channel_map,
    ]
    result = _run_cli(cmd)
    assert result.returncode == 0, result.stderr

    run_dir = Path(result.stdout.strip())
    assert (run_dir / "manifest.json").exists()
    assert (run_dir / "qc.json").exists()
    assert (run_dir / "report.html").exists()
    assert (run_dir / "overlays").exists()
    assert not (run_dir / "results.xlsx").exists()



def test_cli_instructor_override_can_export_excel(tmp_path: Path) -> None:
    img = make_synthetic_yxc(n_nuclei=2)
    inp = tmp_path / "img2.ome.tif"
    write_ome_tiff(inp, img, pixel_size_um=0.5)

    out_root = tmp_path / "runs"
    channel_map = json.dumps({"NUCLEUS": 0, "MARKER_1": 1})

    cmd = [
        sys.executable,
        "-m",
        "protocolquant.cli",
        "run",
        "--protocol",
        "protocols/nuclei_count_intensity_2d_v1.yaml",
        "--input",
        str(inp),
        "--role",
        "instructor",
        "--policy",
        "configs/lab_policy.yaml",
        "--out",
        str(out_root),
        "--accept-preview",
        "--preview-count",
        "1",
        "--channel-map",
        channel_map,
        "--override-fail",
        "--override-reason",
        "authorised", 
    ]
    result = _run_cli(cmd)
    assert result.returncode == 0, result.stderr

    run_dir = Path(result.stdout.strip())
    excel_path = run_dir / "results.xlsx"
    assert excel_path.exists()
    wb = load_workbook(excel_path)
    expected = {"RunSummary", "ImageSummary", "QC", "Provenance", "NucleiTable"}
    assert expected.issubset(set(wb.sheetnames))
